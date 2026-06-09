import argparse
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text or text.lower() == "none":
            return None
        return text
    return value


def clean_int(value, default=0):
    value = clean(value)
    if value is None:
        return default
    return int(float(value))


def clean_float(value):
    value = clean(value)
    if value is None:
        return None
    return float(value)


def normalized(value):
    text = clean(value) or ""
    return "".join(ch.lower() for ch in str(text) if ch.isalnum())


def load_fighter_rows(workbook_path):
    wb = load_workbook(workbook_path, data_only=True)
    fighters_ws = wb["fighters"]
    profiles_ws = wb["fighter_profiles"]
    stats_ws = wb["fighter_stats"]
    events_ws = wb["events_log"]

    fighters = []
    for row_idx in range(2, fighters_ws.max_row + 1):
        spreadsheet_id = clean(fighters_ws.cell(row_idx, 1).value)
        name = clean(fighters_ws.cell(row_idx, 2).value)
        if not spreadsheet_id and not name:
            continue
        fighters.append(
            {
                "spreadsheet_id": spreadsheet_id,
                "name": name,
                "tier": clean(fighters_ws.cell(row_idx, 3).value) or "Tier 3",
                "current_cost": clean_int(fighters_ws.cell(row_idx, 4).value, 0),
                "next_season_cost": clean_int(fighters_ws.cell(row_idx, 5).value, clean_int(fighters_ws.cell(row_idx, 4).value, 0)),
                "height": clean_float(fighters_ws.cell(row_idx, 6).value),
                "weight": clean_float(fighters_ws.cell(row_idx, 7).value),
                "active": 0 if clean(fighters_ws.cell(row_idx, 8).value) in {False, "FALSE", "false", 0} else 1,
                "image_url": clean(fighters_ws.cell(row_idx, 9).value),
                "notes": clean(fighters_ws.cell(row_idx, 10).value) or "",
            }
        )

    fighters_by_id = {row["spreadsheet_id"]: row for row in fighters if row["spreadsheet_id"]}
    fighters_by_name = {row["name"]: row for row in fighters if row["name"]}
    fighters_by_norm = {normalized(row["name"]): row for row in fighters if row["name"]}

    def resolve_roster_row(profile_id, profile_name):
        if profile_id and profile_id in fighters_by_id:
            return fighters_by_id[profile_id]
        if profile_name and profile_name in fighters_by_name:
            return fighters_by_name[profile_name]
        norm_candidates = [normalized(profile_id), normalized(profile_name)]
        for candidate in norm_candidates:
            if candidate and candidate in fighters_by_norm:
                return fighters_by_norm[candidate]
        for candidate in norm_candidates:
            if not candidate:
                continue
            prefix_matches = [row for row in fighters if normalized(row["spreadsheet_id"]) == candidate or normalized(row["name"]).startswith(candidate) or candidate.startswith(normalized(row["name"]))]
            if len(prefix_matches) == 1:
                return prefix_matches[0]
        return None

    for row_idx in range(2, profiles_ws.max_row + 1):
        profile_id = clean(profiles_ws.cell(row_idx, 1).value)
        profile_name = clean(profiles_ws.cell(row_idx, 2).value)
        if not profile_id and not profile_name:
            continue
        roster_row = resolve_roster_row(profile_id, profile_name)
        if not roster_row:
            continue
        roster_row.update(
            {
                "age": clean_int(profiles_ws.cell(row_idx, 3).value, 0) or None,
                "height": clean_float(profiles_ws.cell(row_idx, 4).value) or roster_row.get("height"),
                "weight": clean_float(profiles_ws.cell(row_idx, 5).value) or roster_row.get("weight"),
                "start_year": clean_int(profiles_ws.cell(row_idx, 6).value, 0) or None,
                "nickname": clean(profiles_ws.cell(row_idx, 7).value),
                "fighting_style": clean(profiles_ws.cell(row_idx, 8).value),
                "preferred_role": clean(profiles_ws.cell(row_idx, 9).value),
                "reputation": clean(profiles_ws.cell(row_idx, 10).value),
                "bio": clean(profiles_ws.cell(row_idx, 11).value),
                "fantasy_insight": clean(profiles_ws.cell(row_idx, 12).value),
                "image_url": clean(profiles_ws.cell(row_idx, 13).value) or roster_row.get("image_url"),
            }
        )

    stat_rows = {}
    for row_idx in range(2, stats_ws.max_row + 1):
        fighter_id = clean(stats_ws.cell(row_idx, 1).value)
        fighter_name = clean(stats_ws.cell(row_idx, 2).value)
        if not fighter_id and not fighter_name:
            continue
        roster_row = resolve_roster_row(fighter_id, fighter_name)
        if not roster_row:
            continue
        stat_rows[roster_row["spreadsheet_id"]] = {
            "training": clean_int(stats_ws.cell(row_idx, 6).value, 0),
            "competitions": clean_int(stats_ws.cell(row_idx, 7).value, 0),
            "rounds_fought": clean_int(stats_ws.cell(row_idx, 8).value, 0),
            "support": clean_int(stats_ws.cell(row_idx, 9).value, 0),
            "special_awards": clean_int(stats_ws.cell(row_idx, 10).value, 0),
            "gold_medals": clean_int(stats_ws.cell(row_idx, 11).value, 0),
            "silver_medals": clean_int(stats_ws.cell(row_idx, 12).value, 0),
            "bronze_medals": clean_int(stats_ws.cell(row_idx, 13).value, 0),
            "kills": clean_int(stats_ws.cell(row_idx, 14).value, 0),
            "assists": clean_int(stats_ws.cell(row_idx, 15).value, 0),
            "deaths": clean_int(stats_ws.cell(row_idx, 16).value, 0),
            "sit_downs": clean_int(stats_ws.cell(row_idx, 17).value, 0),
            "yellow_cards": clean_int(stats_ws.cell(row_idx, 18).value, 0),
            "red_cards": clean_int(stats_ws.cell(row_idx, 19).value, 0),
            "ownership_percent": float(clean(stats_ws.cell(row_idx, 29).value) or 0),
        }

    event_rows = []
    for row_idx in range(2, events_ws.max_row + 1):
        fighter_id = clean(events_ws.cell(row_idx, 4).value)
        fighter_name = clean(events_ws.cell(row_idx, 5).value)
        event_name = clean(events_ws.cell(row_idx, 3).value)
        event_date = clean(events_ws.cell(row_idx, 2).value)
        roster_row = resolve_roster_row(fighter_id, fighter_name)
        if not roster_row or not event_name or not event_date:
            continue
        if hasattr(event_date, "strftime"):
            event_date = event_date.strftime("%Y-%m-%d")
        event_rows.append(
            {
                "spreadsheet_id": roster_row["spreadsheet_id"],
                "event_date": str(event_date),
                "event_name": event_name,
                "gold_medals": clean_int(events_ws.cell(row_idx, 6).value, 0),
                "silver_medals": clean_int(events_ws.cell(row_idx, 7).value, 0),
                "bronze_medals": clean_int(events_ws.cell(row_idx, 8).value, 0),
                "kills": clean_int(events_ws.cell(row_idx, 9).value, 0),
                "deaths": clean_int(events_ws.cell(row_idx, 10).value, 0),
                "sit_downs": clean_int(events_ws.cell(row_idx, 11).value, 0),
                "yellow_cards": clean_int(events_ws.cell(row_idx, 12).value, 0),
                "red_cards": clean_int(events_ws.cell(row_idx, 13).value, 0),
            }
        )

    return fighters, stat_rows, event_rows


def sync_workbook(workbook_path, db_path):
    fighters, stat_rows, event_rows = load_fighter_rows(workbook_path)
    imported_at = datetime.now(timezone.utc).isoformat(timespec="seconds")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        fighter_id_by_sheet_id = {}
        for fighter in fighters:
            existing = None
            if fighter["spreadsheet_id"]:
                existing = conn.execute("SELECT id FROM fighters WHERE spreadsheet_id=?", (fighter["spreadsheet_id"],)).fetchone()
            if not existing and fighter["name"]:
                existing = conn.execute("SELECT id FROM fighters WHERE name=?", (fighter["name"],)).fetchone()
            if existing:
                fighter_db_id = existing["id"]
                conn.execute(
                    """
                    UPDATE fighters
                    SET spreadsheet_id=?,
                        name=?,
                        tier=?,
                        height=?,
                        weight=?,
                        age=?,
                        active=?,
                        start_year=?,
                        current_cost=?,
                        next_season_cost=?,
                        notes=?,
                        nickname=COALESCE(?, nickname),
                        fighting_style=COALESCE(?, fighting_style),
                        preferred_role=COALESCE(?, preferred_role),
                        reputation=COALESCE(?, reputation),
                        bio=COALESCE(?, bio),
                        fantasy_insight=COALESCE(?, fantasy_insight),
                        image_url=COALESCE(?, image_url)
                    WHERE id=?
                    """,
                    (
                        fighter["spreadsheet_id"],
                        fighter["name"],
                        fighter["tier"],
                        fighter.get("height"),
                        fighter.get("weight"),
                        fighter.get("age"),
                        fighter.get("active", 1),
                        fighter.get("start_year"),
                        fighter["current_cost"],
                        fighter["next_season_cost"],
                        fighter.get("notes", ""),
                        fighter.get("nickname"),
                        fighter.get("fighting_style"),
                        fighter.get("preferred_role"),
                        fighter.get("reputation"),
                        fighter.get("bio"),
                        fighter.get("fantasy_insight"),
                        fighter.get("image_url"),
                        fighter_db_id,
                    ),
                )
            else:
                cursor = conn.execute(
                    """
                    INSERT INTO fighters(
                        spreadsheet_id,name,tier,height,weight,age,active,start_year,current_cost,next_season_cost,
                        notes,nickname,fighting_style,preferred_role,reputation,bio,fantasy_insight,image_url
                    )
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        fighter["spreadsheet_id"],
                        fighter["name"],
                        fighter["tier"],
                        fighter.get("height"),
                        fighter.get("weight"),
                        fighter.get("age"),
                        fighter.get("active", 1),
                        fighter.get("start_year"),
                        fighter["current_cost"],
                        fighter["next_season_cost"],
                        fighter.get("notes", ""),
                        fighter.get("nickname"),
                        fighter.get("fighting_style"),
                        fighter.get("preferred_role"),
                        fighter.get("reputation"),
                        fighter.get("bio"),
                        fighter.get("fantasy_insight"),
                        fighter.get("image_url"),
                    ),
                )
                fighter_db_id = cursor.lastrowid
            fighter_id_by_sheet_id[fighter["spreadsheet_id"]] = fighter_db_id

        conn.execute("DELETE FROM baseline_stats")
        conn.execute("DELETE FROM attendance_scores")
        conn.execute("DELETE FROM fighter_import_totals")
        conn.execute("DELETE FROM event_results")

        for fighter in fighters:
            fighter_db_id = fighter_id_by_sheet_id[fighter["spreadsheet_id"]]
            stats = stat_rows.get(fighter["spreadsheet_id"], {})
            conn.execute(
                "INSERT INTO baseline_stats(fighter_id,training,support) VALUES(?,?,?)",
                (fighter_db_id, stats.get("training", 0), stats.get("support", 0)),
            )
            conn.execute(
                """
                INSERT INTO fighter_import_totals(
                    fighter_id,training,competitions,rounds_fought,support,special_awards,
                    gold_medals,silver_medals,bronze_medals,kills,assists,deaths,sit_downs,
                    yellow_cards,red_cards,ownership_percent,imported_at
                )
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    fighter_db_id,
                    stats.get("training", 0),
                    stats.get("competitions", 0),
                    stats.get("rounds_fought", 0),
                    stats.get("support", 0),
                    stats.get("special_awards", 0),
                    stats.get("gold_medals", 0),
                    stats.get("silver_medals", 0),
                    stats.get("bronze_medals", 0),
                    stats.get("kills", 0),
                    stats.get("assists", 0),
                    stats.get("deaths", 0),
                    stats.get("sit_downs", 0),
                    stats.get("yellow_cards", 0),
                    stats.get("red_cards", 0),
                    stats.get("ownership_percent", 0.0),
                    imported_at,
                ),
            )

        for event in event_rows:
            fighter_db_id = fighter_id_by_sheet_id.get(event["spreadsheet_id"])
            if not fighter_db_id:
                continue
            conn.execute(
                """
                INSERT INTO event_results(
                    event_date,event_name,fighter_id,rounds_fought,special_awards,gold_medals,silver_medals,bronze_medals,kills,assists,deaths,sit_downs,yellow_cards,red_cards
                )
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    event["event_date"],
                    event["event_name"],
                    fighter_db_id,
                    0,
                    0,
                    event["gold_medals"],
                    event["silver_medals"],
                    event["bronze_medals"],
                    event["kills"],
                    0,
                    event["deaths"],
                    event["sit_downs"],
                    event["yellow_cards"],
                    event["red_cards"],
                ),
            )

        conn.commit()
        print(f"Synchronized {len(fighters)} fighters and {len(event_rows)} event rows from {workbook_path}")
    finally:
        conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--db", required=True)
    args = parser.parse_args()
    sync_workbook(Path(args.workbook), Path(args.db))


if __name__ == "__main__":
    main()
